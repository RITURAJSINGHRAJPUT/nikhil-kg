import openpyxl
import json
import re

def parse_bakery_machines(wb_path):
    wb = openpyxl.load_workbook(wb_path)
    ws = wb['Bakery Machine List']
    
    machines = []
    # Row 1 is title: BAKERY MACHINE LIST SURAT
    # Row 2 is headers: Sr. | Machine / Item | Qty | Pic | Equipment Code | Brand | Capacity | Purchase Date | QR Code
    headers = []
    for cell in ws[2]:
        headers.append(cell.value)
    
    print("Bakery Headers:", headers)
    
    for row_idx in range(3, ws.max_row + 1):
        row = [cell.value for cell in ws[row_idx]]
        if not row or not row[0] or str(row[0]).strip() == "":
            continue
        
        # Check if row is headers or section title
        if "Machine" in str(row[1]) or "Sr." in str(row[0]):
            continue
            
        code = str(row[4]).strip() if row[4] else f"SUR-BKY-{row_idx-2:03d}"
        
        item = {
            "id": code,
            "category": "Bakery Machine List",
            "name": str(row[1]).strip() if row[1] else "",
            "qty": str(row[2]).strip() if row[2] else "1",
            "brand": str(row[5]).strip() if row[5] else "-",
            "capacity": str(row[6]).strip() if row[6] else "-",
            "purchase_date": str(row[7]).strip() if row[7] else "-",
            "details": f"Equipment Code: {code}. Brand: {row[5] or '-'}. Capacity: {row[6] or '-'}. Purchase Date: {row[7] or '-'}"
        }
        machines.append(item)
    return machines

def parse_prep_kitchen_sheet1(wb_path):
    # Sheet1 in Bakery_Machine_list_with_QR_5.xlsx has contact numbers and use/application
    wb = openpyxl.load_workbook(wb_path)
    ws = wb['Sheet1']
    
    machines = []
    # Row 1: Prep Kitchen Machine Litst Surat
    # Row 2: Sr No | Product Name | Qty | Use/Application | Brand | contact nomber
    for row_idx in range(3, ws.max_row + 1):
        row = [cell.value for cell in ws[row_idx]]
        if not row or not row[0] or str(row[0]).strip() == "":
            continue
            
        name = str(row[1]).strip() if row[1] else ""
        if "Product Name" in name:
            continue
            
        code = f"SUR-PRP-{row_idx-2:03d}"
        
        item = {
            "id": code,
            "category": "Prep Kitchen",
            "name": name,
            "qty": str(row[2]).strip() if row[2] else "1",
            "brand": str(row[4]).strip() if row[4] else "-",
            "use": str(row[3]).strip() if row[3] else "-",
            "contact": str(row[5]).strip() if row[5] else "-",
            "details": f"Use: {row[3] or '-'}. Brand: {row[4] or '-'}. Contact: {row[5] or '-'}"
        }
        machines.append(item)
    return machines

def parse_new_machines(wb_path):
    # Let's extract new machines from "All machine Prep kitchen & ODC"
    # Starting from "New Machine Available At Prep Kitchen - 2026"
    wb = openpyxl.load_workbook(wb_path)
    ws = wb['All machine Prep kitchen & ODC']
    
    machines = []
    start_parsing = False
    new_idx = 1
    
    for row_idx in range(1, ws.max_row + 1):
        row = [cell.value for cell in ws[row_idx]]
        row_str = " ".join([str(c) for c in row if c is not None])
        
        if "New Machine Available At Prep Kitchen" in row_str:
            start_parsing = True
            continue
            
        if start_parsing:
            # check if we reached next section "Order Done In Process"
            if "Order Done" in row_str or "Pending Machine" in row_str or "SODHELA" in row_str:
                # Keep parsing, or stop if we hit other boundaries. Let's see what is under New Machine Available.
                pass
            
            # A row with numeric Sr No
            val_0 = str(row[0]).strip() if row[0] is not None else ""
            if val_0.isdigit():
                name = str(row[1]).strip() if row[1] else ""
                qty = str(row[2]).strip() if row[2] else "1"
                remark = str(row[5]).strip() if row[5] else "-"
                
                code = f"SUR-NEW-{new_idx:03d}"
                new_idx += 1
                
                item = {
                    "id": code,
                    "category": "New Machines (2026)",
                    "name": name,
                    "qty": qty,
                    "remark": remark,
                    "brand": "-",
                    "details": f"Remark: {remark}"
                }
                machines.append(item)
                
            # If we see empty rows or next big titles we can continue or stop depending on structure
            # Let's look at the rows under this section.
    return machines

def parse_rd_kitchen(wb_path):
    wb = openpyxl.load_workbook(wb_path)
    ws = wb['R&D Kitchen معدات']
    
    machines = []
    for row_idx in range(3, ws.max_row + 1):
        row = [cell.value for cell in ws[row_idx]]
        if not row or not row[0] or str(row[0]).strip() == "":
            continue
            
        name = str(row[1]).strip() if row[1] else ""
        if "Product Name" in name:
            continue
            
        code = f"SUR-RD-{row_idx-2:03d}"
        
        item = {
            "id": code,
            "category": "R&D Kitchen",
            "name": name,
            "qty": str(row[2]).strip() if row[2] else "1",
            "use": str(row[3]).strip() if row[3] else "-",
            "brand": str(row[4]).strip() if row[4] else "-",
            "contact": str(row[5]).strip() if row[5] else "-",
            "details": f"Use: {row[3] or '-'}. Brand: {row[4] or '-'}. Contact: {row[5] or '-'}"
        }
        machines.append(item)
    return machines

if __name__ == '__main__':
    wb5 = 'Bakery_Machine_list_with_QR_5.xlsx'
    
    bakery = parse_bakery_machines(wb5)
    prep = parse_prep_kitchen_sheet1(wb5)
    new_m = parse_new_machines(wb5)
    rd = parse_rd_kitchen(wb5)
    
    all_machines = bakery + prep + new_m + rd
    
    print(f"Loaded: {len(bakery)} Bakery, {len(prep)} Prep, {len(new_m)} New, {len(rd)} R&D. Total: {len(all_machines)}")
    
    with open('machines.json', 'w', encoding='utf-8') as f:
        json.dump(all_machines, f, indent=2, ensure_ascii=False)
    
    print("Exported machines.json successfully!")
